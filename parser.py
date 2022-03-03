#!/usr/bin/python3

import sys, os, re, datetime, pyairtable, logger
import xml.etree.ElementTree as ET
from dotenv import load_dotenv, dotenv_values
from openpyxl import load_workbook


class InputFieldParser:
    """
    encapsulates a rule in the parser xml document. Each rule defines
    a thing to look for in the main document - usually a document-level
    parameter, or the beginning of a table. For tables, the table is handed
    off to a specialised table parser, which returns a list of items
    """
    def __init__(self, tree):
        # initialize variables
        self.value = ''
        self.tree = tree
        self.triggers = []
        self.cells = []
        self.checkfunctions = []
        # load triggers, cells, actions etc
        self.name = tree.find('name').text
        outputfield = tree.find('outputfield')
        # get the output field name
        if outputfield != None:
            self.outputfield = outputfield.text

        for trigger in tree.findall('trigger'):
            self.triggers.append(trigger)
        for cell in tree.findall('cell'):
            self.addCell(cell)
        for checkfunction in tree.findall('checkfunction'):
            self.addCheckFunc(checkfunction)
        for action in tree.findall('action'):
            self.addAction(action)

    def addCheckFunc(self, tree):
        """ add a check function based on data in the xml """
        funcname = tree.text
        checktype = tree.attrib['type']
        self.checkfunctions.append(tree)
        pass

    def runCheckFunctions(self, pofield):
        """ run all of the check functions stored in this parser.
        if ANY ONE of them is True, output 'True'
        concatenate all the report strings together, if there are any, 
        and output that. 
        The output boolean and string are stored in the pofield object
        """ 
        outputstring = ''
        outputflags= []
        for checkfunc in self.checkfunctions:
            checktype = checkfunc.attrib['type']
            if checktype == 'regex': #for regex checkfuncs..
                # get the field 'checkstring' etc - if not present, default to empty string
                checkstring = getTreeText(checkfunc, 'checkstring', '')
                nocheckstring = getTreeText(checkfunc, 'nocheckstring', '')

                # checkif - does the function set check flag if there's a match, or if there is no match?
                # can be True (match) or False (nomatch). 
                checkif = getTreeText(checkfunc, 'checkif', 'match').lower()
                if checkif == 'match':
                    checkif = True
                else:
                    checkif = False

                try:
                    # see if whoever input the function into the xml provided a regex. If not, 
                    # raise an exception.
                    regex = checkfunc.find('regex').text
                    string = pofield.value 
                except AttributeError:
                    print("No regex defined!")
                else:
                    if re.match(regex, string):
                        check = checkif
                    else:
                        check = not checkif

                    if check:
                        outputstring += checkstring
                        outputflags.append(True)
                    else: 
                        outputstring += nocheckstring
                        outputflags.append(False)
                    outputstring += ' '

        if outputflags == []:
            outputflags = [False] #defaults to True (CHECK)

        # collapse the list of output flags together. if any are set to True, 
        # the whole thing will be true ie if anything needs to be CHECKED, 
        # the whole thing will be set to check.
        pofield.checkFlag = any(outputflags) 
        # print(outputstring)
        pofield.checkString = outputstring
        return pofield

    def addAction(self, treeField):
        """
        add an action based on the data in the xml. 
        can be just a direct output, or a more complex function.
        """
        self.actiontype = treeField.text
        # if the action is 'Parse Table', create a TableParser to 
        # encapsulate the table-specific rules. 
        # We always initialize the rules objects FIRST before doing any parsing.
        if self.actiontype == 'Parse Table':
            self.tableparser = TableParser(tree, data_rows)
        # print(self.actiontype)

    def addCell(self, cellObj):
        """
        the 'cell' variable stores a relative vector pointing toward where the 
        data in the spreadsheet is, from the location of the header
        """
        processedList = cellObj.text.replace(' ', '').split(',')
        try: 
            xcoord = int(processedList[0])
        except ValueError:
            print("Invalid cell definition: ", cellObj.text)
        try: 
            ycoord = int(processedList[1])
        except ValueError:
            print("Invalid cell definition: ", cellObj.text)
        cellCoords = (xcoord, ycoord)
        self.cells.append(cellCoords)

    def matchAndDoAction(self, string, data_rows, coordinateTuple):
        """
        matches the input string against all triggers. If there is a match,  
        chooses an action to take based on the 'actiontype' variable
        and executes it. Returns either a POField or, in the case of a table,
        a list of POItems (which contain POFields)
        
        """
        output = None
        if self.matchTrigger(string):
            actionname = self.actiontype
            if actionname == 'Direct Output': 
                actionname = 'getCells'
            elif actionname == 'Parse Table':
                actionname = 'parseTable'
            action = getattr(self, actionname) #find the method of this class that matches the 'actiontype' string
            output = action(self.tree, data_rows, coordinateTuple)
            # check output and store the result in the POField's checkFlag and checkString
            if output.__class__.__name__ == 'POField':
                self.runCheckFunctions(output)
            return output

    def parseTable(self, tree, data_rows, coordinateTuple):
        """ just runs the parse method of the table parser and returns the output """
        output = self.tableparser.parse(coordinateTuple)
        return output

    def matchTrigger(self, string):
        """ checks every trigger and returns a boolean for match """ 
        for trigger in self.triggers:
            txt = trigger.text
            trigtype = trigger.attrib['type']
            if trigtype == 'raw':
                if txt == string:
                    return True
                else:
                    return False
            elif trigtype == 'regex':
                # print(txt, string)
                if re.match(txt, str(string)):
                   return True
                else: 
                    return False

    def getCells(self, tree,  data_rows, coordinateTuple):
        """
        reads all the items in self.cells, and extracts the data from those cells,
        offset from the start cell (coordinateTuple).
        Returns a POField object.
        """
        xcoord = coordinateTuple[0]
        ycoord = coordinateTuple[1]
        output = ''
        for item in self.cells:
            xoffset = item[0]
            yoffset = item[1]
            # y first, then x. Just a quirk of how the spreadsheet is stored. 
            targetcell = data_rows[ycoord+yoffset][xcoord+xoffset]
            if targetcell: 
                if targetcell.__class__.__name__ == 'datetime':
                    output += targetcell.strftime('%Y-%m-%d')
                else:
                    output += targetcell
            output += ' '
        return POField(self.outputfield, output)



class POParser:
    """
    class for parsing PO-level stuff. encapsulates functions performed on the 
    document level.
    """

    def __init__(self, tree):
        """ reads an XML tree on creation and creates a list of InputFieldParsers,
        which store parsing rules """
        self.tree = tree
        self.inputfields = []
        self.outputfields = []
        # get input fields from the tree
        inputs = self.tree.find('inputs')
        outputs = self.tree.find('outputs')
        for item in inputs.findall('field'): #find all fields
            try: 
                fieldname = item.find('name')
            except: 
                print("Can't find item name.")
            else: # make a new InputFieldParser object 
                self.inputfields.append(InputFieldParser(item))
        # populate the list of output fields. These will go into the final PO object.
        for item in outputs.iter('name'):
            self.outputfields.append(item.text)

    def parse(self, data_rows):
        """ 
        reads a spreadsheet in list form and returns a PO object 
        by applying stored rules. 
        """ 
        # for every cell in the sheet, addressed by row and col (x and y)... 
        po_object = PO()
        for rownum, row in enumerate(data_rows):
            for colnum, cell in enumerate(row):
                # if the cell isn't empty, apply every single field's 'matchAndGet' function to the cell to see which is a match
                # inefficient but this scraper doesn't have to be quick. 
                if cell != None:
                    for field in self.inputfields:
                        output = field.matchAndDoAction(cell, data_rows, (colnum, rownum))
                        # output = field.matchAndGet(cell, data_rows, (colnum, rownum))  
                        if output:
                            po_object.addItem(output)

    def funcAllFields(self, func):
        for item in self.inputfields: 
            func(item)

class TableParser:
    """ called by InputFieldParser """
    def __init__(self, tree, data_rows):
        self.tree = tree
        self.rawtable = data_rows
        self.tablefields = []
        for tablefield in tree.findall('tablefield'):
            self.tablefields.append(tablefield.text)
        # print(self.headers)

    def getItemRowNumbers(self, coordinateTuple):
        rowNumbers = []
        ycoord = coordinateTuple[1]
        numRows = len(self.rawtable)
        for rownum in range(ycoord+1, numRows-1):
            if self.rawtable[rownum][0] != None:
                rowNumbers.append(rownum)
            elif self.rawtable[rownum+1][0] == None:
                break
        return rowNumbers

    def getHeadersAndColumnNumbers(self, coordinateTuple):
        colNumbers = []
        xcoord = coordinateTuple[0]
        ycoord = coordinateTuple[1]
        numCols = len(self.rawtable[0])
        for colnum in range(0, numCols - 1):
            header = self.rawtable[ycoord][colnum]
            if colnum == 0: #get the first column header and store as the item field title
                itemHeader = header
            if header != None:
                colNumbers.append((header, colnum))
        return colNumbers, itemHeader

    def parse(self, coordinateTuple):
        """
        returns a list of POItems
        """
        xcoord = coordinateTuple[0]
        ycoord = coordinateTuple[1]
        self.tablestartcoord = (xcoord, ycoord)
        self.headers = [header for header in self.rawtable[ycoord] if header != None]
        rowNumbers = self.getItemRowNumbers(self.tablestartcoord)
        colNumbers, itemHeader = self.getHeadersAndColumnNumbers(self.tablestartcoord)
        output = []
        for row in rowNumbers:
            po_item = POItem()
            for header, column in colNumbers: 
                # make a new POField, with the header and value from the table
                field = POField(header, self.rawtable[row][column])
                #add the field to the growing POItem
                po_item.addField(field)
            output.append(po_item)
                

        return output


class PO:
    def __init__(self):
        self.items = [] 

    def addItem(self, item):
        self.items.append(item)
        # print(item.__class__)
        if item.__class__.__name__ == 'list':
            for poitem in item:
                poitem.printVal()
        else:
            item.printVal()

class POItem:
    def __init__(self):
        self.fields = {}

    def addField(self, pofield):
        self.fields[pofield.header] = pofield

    def printVal(self):
        outputstring = ''
        for key in self.fields:
            outputstring += self.fields[key].header
            outputstring += ': '
            outputstring += self.fields[key].getValString() + ' | '
        print(outputstring)

    def getByHeader(self, header):
        return self.fields[header].value

class POField:
    """ 
    class to store data for a single PO field (as on the excel) 
    ideally, will only be a data struct and not hold a lot of functions
    """ 
    def __init__(self, header, value):
        self.value = value
        self.header = header
        self.checkFlag = False
        self.checkString = ''

    def updateVal(self, newVal):
        self.value = newVal

    def getValString(self):
        return str(self.value)

    def printVal(self):
        print(self.header, ': ', self.value, self.checkFlag, self.checkString)


def getTreeText(tree, fieldname, ifCantFind):
    text = tree.find(fieldname)
    if text != None:
        text = text.text
    else: 
        text = ifCantFind
    return text

if __name__ == '__main__':

    # Load environment and initialize airtable interface
    load_dotenv()
    key = os.environ['AIRTABLEKEY']
    baseid = os.environ['AIRTABLEBASEID']
    remote_table = pyairtable.Table(key, baseid, 'testtable')

    # read excel file
    filepath = sys.argv[1]
    excelfile = load_workbook(filename = filepath, data_only=True)
    sheetnames = excelfile.sheetnames
    data_rows = [] 
    for name in sheetnames: 
        print('Parsing sheet:', name)
        sheet = excelfile[name]
        for row in sheet.iter_rows():
            rowvals = [x.value for x in row]
            data_rows.append(rowvals)

    # Parse the xml document for rules

    tree = ET.parse('data.xml')
    parser = POParser(tree)
    root = parser.tree.getroot()


    # parse the excel data
    parser.parse(data_rows)
    print(parser.outputfields)
    # print(data_rows)


    # parser.funcAllFields(lambda x: print(x.name, x.cells))



