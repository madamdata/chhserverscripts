#!/usr/bin/python3

import sys, os, re, datetime, pyairtable
import xml.etree.ElementTree as ET
from dotenv import load_dotenv, dotenv_values
from openpyxl import load_workbook


class InputFieldParser:
    def __init__(self, tree):
        # initialize variables
        self.value = ''
        self.tree = tree
        self.triggers = []
        self.cells = []
        # load triggers, cells, actions etc
        self.name = tree.find('name').text
        outputfield = tree.find('outputfield')
        # print(outputfield.text)
        if outputfield != None:
            self.outputfield = outputfield.text
        else:
            self.outputfield = None
        for trigger in tree.findall('trigger'):
            self.triggers.append(trigger)
        for cell in tree.findall('cell'):
            self.addCell(cell)
        for action in tree.findall('action'):
            self.addAction(action)

    def addAction(self, treeField):
        """
        add an action based on the data in the xml. 
        can be just a direct output, or a more complex function.
        """
        self.actiontype = treeField.text
        # print(self.actiontype)

    def addCell(self, cellObj):
        """
        the 'cell' variable stores a relative vector pointing toward where the 
        data in the spreadsheet is, from the location of the header
        """
        processedList = cellObj.text.replace(' ', '').split(',')
        try: 
            xcoord = int(processedList[0])
        except ValueError:
            print("Invalid cell definition: ", cellObj.text)
        try: 
            ycoord = int(processedList[1])
        except ValueError:
            print("Invalid cell definition: ", cellObj.text)
        cellCoords = (xcoord, ycoord)
        self.cells.append(cellCoords)

    def matchAndDoAction(self, string, data_rows, coordinateTuple):
        """
        matches the input string against all triggers. If there is a match,  
        chooses an action to take based on the 'actiontype' variable
        and executes it. Returns a dictionary of key-value pairs corresponding to output fields
        and data in the PO object.
        
        """
        output = None
        if self.matchTrigger(string):
            actionname = self.actiontype
            if actionname == 'Direct Output': 
                actionname = 'getCells'
            elif actionname == 'Parse Table':
                actionname = 'createTableParser'
            action = getattr(self, actionname)
            output = action(self.tree, data_rows, coordinateTuple)
            return output

    def createTableParser(self, tree, data_rows, coordinateTuple):
        tableparser = TableParser(tree, data_rows, coordinateTuple) 
        output = tableparser.parse()
        return output

    def matchTrigger(self, string):
        """ checks every trigger and returns a boolean for match """ 
        for trigger in self.triggers:
            txt = trigger.text
            trigtype = trigger.attrib['type']
            if trigtype == 'raw':
                if txt == string:
                    return True
                else:
                    return False
            elif trigtype == 'regex':
                # print(txt, string)
                if re.match(txt, str(string)):
                   return True
                else: 
                    return False

    def getCells(self, tree,  data_rows, coordinateTuple):
        """
        reads all the items in self.cells, and extracts the data from those cells,
        offset from the start cell (coordinateTuple).
        Returns a dictionary with --> 
        one key: the name of this parser, 
        and one value: all the contents of those cells concatenated.
        """
        xcoord = coordinateTuple[0]
        ycoord = coordinateTuple[1]
        output = ''
        for item in self.cells:
            xoffset = item[0]
            yoffset = item[1]
            # y first, then x. Just a quirk of how the spreadsheet is stored. 
            targetcell = data_rows[ycoord+yoffset][xcoord+xoffset]
            if targetcell: 
                if targetcell.__class__.__name__ == 'datetime':
                    output += targetcell.strftime('%Y-%m-%d')
                else:
                    output += targetcell
            output += ' '
        return {self.outputfield: output}



class POParser:
    """
    class for parsing PO-level stuff. encapsulates functions performed on the 
    document level.
    """

    def __init__(self, tree):
        """ reads an XML tree on creation and creates a list of InputFieldParsers,
        which store parsing rules """
        self.tree = tree
        self.inputfields = []
        self.outputfields = []
        # get input fields from the tree
        inputs = self.tree.find('inputs')
        outputs = self.tree.find('outputs')
        for item in inputs.findall('field'): #find all fields
            try: 
                fieldname = item.find('name')
            except: 
                print("Can't find item name.")
            else: # make a new InputFieldParser object 
                self.inputfields.append(InputFieldParser(item))
        # populate the list of output fields. These will go into the final PO object.
        for item in outputs.iter('name'):
            self.outputfields.append(item.text)

    def parse(self, data_rows):
        """ 
        reads a spreadsheet in list form and returns a PO object 
        by applying stored rules. 
        """ 
        # for every cell in the sheet, addressed by row and col (x and y)... 
        po_object = PO()
        for rownum, row in enumerate(data_rows):
            for colnum, cell in enumerate(row):
                # if the cell isn't empty, apply every single field's 'matchAndGet' function to the cell to see which is a match
                # inefficient but this scraper doesn't have to be quick. 
                if cell != None:
                    for field in self.inputfields:
                        output = field.matchAndDoAction(cell, data_rows, (colnum, rownum))
                        # output = field.matchAndGet(cell, data_rows, (colnum, rownum))  
                        if output:
                            print(output)

    def funcAllFields(self, func):
        for item in self.inputfields: 
            func(item)

class TableParser:
    """ called by InputFieldParser """
    def __init__(self, tree, data_rows, coordinateTuple):
        xcoord = coordinateTuple[0]
        ycoord = coordinateTuple[1]
        self.tablestartcoord = (xcoord, ycoord)
        self.tree = tree
        self.rawtable = data_rows
        self.headers = [header for header in self.rawtable[ycoord] if header != None]
        print(self.headers)

    def getItemRowNumbers(self, coordinateTuple):
        rowNumbers = []
        ycoord = coordinateTuple[1]
        numRows = len(self.rawtable)
        for rownum in range(ycoord+1, numRows-1):
            if self.rawtable[rownum][0] != None:
                rowNumbers.append(rownum)
            elif self.rawtable[rownum+1][0] == None:
                break
        return rowNumbers

    def getHeadersAndColumnNumbers(self, coordinateTuple):
        colNumbers = []
        xcoord = coordinateTuple[0]
        ycoord = coordinateTuple[1]
        numCols = len(self.rawtable[0])
        for colnum in range(0, numCols - 1):
            header = self.rawtable[ycoord][colnum]
            if header != None:
                colNumbers.append((header, colnum))
        return colNumbers

    def parse(self):
        rowNumbers = self.getItemRowNumbers(self.tablestartcoord)
        colNumbers = self.getHeadersAndColumnNumbers(self.tablestartcoord)
        for row in rowNumbers:
            item = []
            for header, column in colNumbers: 
                item.append((header, self.rawtable[row][column]))
            print(item)

        print('rowNumbers', rowNumbers)
        print('colNumbers', colNumbers)
        return {'foo': 3, 'bar': 'xyz'}


class PO:
    def __init__(self):
        self.items = [] 

class POItem:
    def __init__(self):
        self.fields = []
    

def printAll(iterobj, func):
    for item in iterobj: 
        print(func(iterobj[item]))

if __name__ == '__main__':

    # Load environment and initialize airtable interface
    load_dotenv()
    key = os.environ['AIRTABLEKEY']
    baseid = os.environ['AIRTABLEBASEID']
    remote_table = pyairtable.Table(key, baseid, 'testtable')

    # read excel file
    filepath = sys.argv[1]
    excelfile = load_workbook(filename = filepath, data_only=True)
    sheetnames = excelfile.sheetnames
    data_rows = [] 
    for name in sheetnames: 
        print('Parsing sheet:', name)
        sheet = excelfile[name]
        for row in sheet.iter_rows():
            rowvals = [x.value for x in row]
            data_rows.append(rowvals)

    # Parse the xml document for rules

    tree = ET.parse('data.xml')
    parser = POParser(tree)
    root = parser.tree.getroot()


    # parse the excel data
    parser.parse(data_rows)
    print(parser.outputfields)
    # print(data_rows)


    # parser.funcAllFields(lambda x: print(x.name, x.cells))



