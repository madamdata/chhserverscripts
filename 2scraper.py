#!/usr/bin/python3

import parser as poparser
import processor as poprocessor
import sys, os, pyairtable, logger, argparse, datetime
import xml.etree.ElementTree as ET
from dotenv import load_dotenv, dotenv_values
from openpyxl import load_workbook

if __name__ == '__main__':


    # Load environment and initialize airtable interface
    # load_dotenv()
    # key = os.environ['AIRTABLEKEY']
    # baseid = os.environ['AIRTABLEBASEID']
    # remote_table = pyairtable.Table(key, baseid, 'testtable')

    # TEMPORARY FOR WOLTER SCRAPING
    load_dotenv()
    key = os.environ['AIRTABLEKEY']
    baseid = os.environ['AIRTABLEBASEID']
    remote_table = pyairtable.Table(key, baseid, 'Table 1')
    test_table = pyairtable.Table(key, baseid, 'test table')
    datestring = datetime.datetime.now().strftime("%d%b%Y-%H:%M")

    
    # parse script args
    argparser = argparse.ArgumentParser(description="Scraper v2.0")
    argparser.add_argument('-potype')
    argparser.add_argument('-mode', nargs='?', const='dryrun')
    argparser.add_argument('filepath')
    args = argparser.parse_args()
    # print(args.filepath, args.potype)

    filepath = args.filepath
    mode = args.mode
    potype = args.potype
    testupload = False
    realupload = False
    dryrun = False

    if mode == 'dryrun':
        dryrun = True

    if mode == 'testupload':
        testupload = True

    if mode == 'realupload':
        realupload = True

    # read excel file
    filename = os.path.basename(filepath)
    excelfile = load_workbook(filename = filepath, data_only=True)
    sheetnames = excelfile.sheetnames
    sheetFileMismatch = False
    sheetdata = []

    print(' >>>>>>> BEGIN ', filename, ' >>>>>>>>>\n')
    for name in sheetnames: 
        data_rows = [] 
        print(datestring, '- Parsing sheet:', name)
        if name not in filename:
            print('sheet name not the same as filename!!')
            sheetFileMismatch = True
        sheet = excelfile[name]
        for row in sheet.iter_rows():
            rowvals = [x.value for x in row]
            newrowvals = []
            # print(rowvals)
            for item in rowvals:
                newitem = item
                if item == ' ':
                    newitem = None
                newrowvals.append(newitem)

            data_rows.append(newrowvals)
        sheetdata.append(data_rows)

    # Parse the xml document for rules

    if potype == 'wolter': 
        parsertree = ET.parse('/home/chh/mail/chhserverscripts/rules-parser-wolter.xml')
    elif potype == 'rosenberg':
        parsertree = ET.parse('/home/chh/mail/chhserverscripts/rules-parser.xml')
    parser = poparser.POParser(parsertree)
    parser.filename = filename


    # parse the excel data from THE FIRST SHEET ONLY and get PO object
    po = parser.parse(sheetdata[0])
    if sheetFileMismatch:       
        po.addExtraCheckStrings('sheetFileMismatch')
    # check for multiple sheets
    if len(sheetnames) > 1:
        multiple_sheets = True
        po.addExtraCheckStrings('MULTIPLE SHEETS IN THIS FILE')

    po.printAll()
    po.collectCheckStrings()
    po.summarizeCheckStrings()

    #PROCESSOR

    rosenberg_processortree = ET.parse('/home/chh/mail/chhserverscripts/rules-processor-rosenberg.xml')
    wolter_processortree = ET.parse('/home/chh/mail/chhserverscripts/rules-processor-wolter.xml')
    if potype == 'wolter': 
        processor = poprocessor.POProcessor(wolter_processortree)
    elif potype == 'rosenberg':
        processor = poprocessor.POProcessor(rosenberg_processortree)

    nodenetwork = processor.parse(po)
    # nodenetwork.listNodes()
    if testupload:
        processor.upload('upload group 1', test_table, printout=True)

    if realupload:
        processor.upload('upload group 1', remote_table, printout=False)

    if dryrun:
        processor.upload('upload group 1', test_table, printout=True, dryrun=True)

    # nodenetwork.listNodes(nodenames = ['Price per Unit', 's.up'])
    # nodenetwork.listNodes(nodenames = ['Note Raw', 'bracket1st', 'bracket2nd', 'F/B (scraped)'])
    # nodenetwork.listNodes(nodenames=['ITEM', 'modelstringItem', 'MODEL'])
    # nodenetwork.listNodes(nodenames=['ITEM', 'splitA', 'modelstringItem', 'MODEL', 'modelstringExtra'])
    # nodenetwork.listNodes(nodenames=['ITEM', 'detailSplit'])

    print(' >>>>>>> END ', filename, '>>>>>>>\n')


